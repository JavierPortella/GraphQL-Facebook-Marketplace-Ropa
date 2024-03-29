from datetime import datetime, timedelta
from json import loads, JSONDecodeError
from logging import (
    basicConfig,
    CRITICAL,
    ERROR,
    FileHandler,
    getLogger,
    INFO,
    log,
    shutdown,
    StreamHandler,
)
from os import environ, getenv, makedirs, path
from re import sub
from time import localtime, sleep, strftime, time
from traceback import TracebackException

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from pandas import DataFrame
from seleniumwire.webdriver import Chrome, ChromeOptions
from seleniumwire.utils import decode
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    ElementNotInteractableException,
    TimeoutException,
)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.remote_connection import LOGGER as seleniumLogger
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from urllib3.connectionpool import log as urllibLogger
from webdriver_manager.chrome import ChromeDriverManager

CURRENT_DATE = datetime.now().date()


class Errores:
    """Representa a los errores ocurridos durante la ejecución de un scraper

    Attributes:
        errores (dict): Conjunto de datos que contiene toda información de los errores ocurridos durante la ejecución del scraper
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase Errores"""
        self._errores = {
            "Clase": [],
            "Mensaje": [],
            "Linea de Error": [],
            "Codigo Error": [],
            "Publicacion": [],
        }

    @property
    def errores(self):
        """Retorna el valor actual del atributo errores"""
        return self._errores

    def agregar_error(self, error, enlace=None):
        """Agrega la información de un nuevo error al conjunto de datos errores

        Args:
            error (Exception): Error ocurrido durante la ejecución del scraper
            enlace (str, optional): Enlace de la publicación de la página facebook marketplace. Defaults to None.
        """
        log(ERROR, f"Error:\n{error}")
        traceback_error = TracebackException.from_exception(error)
        error_stack = traceback_error.stack[0]
        self._errores["Clase"].append(traceback_error.exc_type)
        self._errores["Mensaje"].append(traceback_error._str)
        self._errores["Linea de Error"].append(error_stack.lineno)
        self._errores["Codigo Error"].append(error_stack.line)
        self._errores["Publicacion"].append(enlace)


class Dataset:
    """Representa al conjunto de datos generado por el scraper

    Attributes:
        dataset (dict): Conjunto de datos que contiene toda información extraída de las publicaciones de la página de facebook marketplace
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase Dataset"""
        self._dataset = {
            "Fecha Extraccion": [],
            "titulo_marketplace": [],
            "tiempo_creacion": [],
            "tipo_delivery": [],
            "descripcion": [],
            "disponible": [],
            "vendido": [],
            "fecha_union_vendedor": [],
            "cantidad": [],
            "precio": [],
            "tipo_moneda": [],
            "amount_with_concurrency": [],
            "latitud": [],
            "longitud": [],
            "locacion": [],
            "locacion_id": [],
            "name_vendedor": [],
            "tipo_vendedor": [],
            "id_vendedor": [],
            "enlace": [],
        }

    @property
    def dataset(self):
        """Retorna el valor actual del diccionario de datos dataset"""
        return self._dataset

    def agregar_data(self, item, fecha_extraccion, enlace):
        """Agrega la información de una publicación al conjunto de datos dataset

        Args:
            item (dict): Conjunto de datos que contiene toda la información de una publicación
            fecha_extraccion (str): Fecha correspondiente a la extracción de todas las publicaciones
            enlace (str): Enlace de la publicación
        """
        self._dataset["Fecha Extraccion"].append(fecha_extraccion)
        self._dataset["titulo_marketplace"].append(
            item.get("marketplace_listing_title")
        )
        self._dataset["tiempo_creacion"].append(item.get("creation_time"))
        try:
            self._dataset["tipo_delivery"].append(item.get("delivery_types")[0])
        except:
            self._dataset["tipo_delivery"].append(None)
        try:
            self._dataset["descripcion"].append(
                item.get("redacted_description").get("text")
            )
        except:
            self._dataset["descripcion"].append(None)
        self._dataset["disponible"].append(item.get("is_live"))
        self._dataset["vendido"].append(item.get("is_sold"))
        try:
            self._dataset["fecha_union_vendedor"].append(
                item.get("marketplace_listing_seller").get("join_time")
            )
        except:
            self._dataset["fecha_union_vendedor"].append(None)
        self._dataset["cantidad"].append(item.get("listing_inventory_type"))
        try:
            listing_price = item.get("listing_price")
            self._dataset["precio"].append(listing_price.get("amount"))
            self._dataset["tipo_moneda"].append(listing_price.get("currency"))
            self._dataset["amount_with_concurrency"].append(
                listing_price.get("amount_with_offset_in_currency")
            )
        except:
            self._dataset["precio"].append(None)
            self._dataset["tipo_moneda"].append(None)
            self._dataset["amount_with_concurrency"].append(None)
        try:
            location = item.get("location")
            self._dataset["latitud"].append(location.get("latitude"))
            self._dataset["longitud"].append(location.get("longitude"))
        except:
            self._dataset["latitud"].append(None)
            self._dataset["longitud"].append(None)
        try:
            self._dataset["locacion"].append(item.get("location_text").get("text"))
        except:
            self._dataset["locacion"].append(None)

        self._dataset["locacion_id"].append(item.get("location_vanity_or_id"))
        try:
            user = item.get("story").get("actors")[0]
            self._dataset["name_vendedor"].append(user.get("name"))
            self._dataset["tipo_vendedor"].append(user.get("__typename"))
            self._dataset["id_vendedor"].append(user.get("id"))
        except:
            self._dataset["name_vendedor"].append(None)
            self._dataset["tipo_vendedor"].append(None)
            self._dataset["id_vendedor"].append(None)
        self._dataset["enlace"].append(enlace)


class Tiempo:
    """Representa al tiempo de ejecución del scraper

    Attributes
        start (float): Hora de inicio de la ejecución del scraper en segundos
        fecha (str): Fecha de extracción de las publicaciones en formato %d/%m/%Y
        hora_inicio (str): Hora de inicio de la ejecución del scraper en formato %H:%M:%S
        hora_fin (str): Hora de término de la ejecución del scraper en formato %H:%M:%S
        cantidad (int): Cantidad de publicaciones extraídas de la página de facebook marketplace
        cantidad_real (int): Cantidad de publicaciones analizadas de la página de facebook marketplace
        tiempo (str): Tiempo de ejecución del scraper en formato %d days, %H:%M:%S
        productos_por_min (float): Cantidad de publicaciones que puede extraer el scraper en un minuto
        productos_por_min_real (float): Cantidad publicaciones que puede analizar el scraper en un minuto
        num_error (int): Cantidad de errores ocurridos durante la ejecución del scraper
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase Tiempo"""
        self._start = time()
        self._fecha = CURRENT_DATE.strftime("%d/%m/%Y")
        self._hora_inicio = strftime("%H:%M:%S", localtime(self._start))
        self._hora_fin = None
        self._cantidad = 0
        self._cantidad_real = 0
        self._tiempo = None
        self._productos_por_min = None
        self._productos_por_min_real = None
        self._num_error = None

    @property
    def cantidad(self):
        """Retorna el valor actual o actualiza el valor del atributo cantidad"""
        return self._cantidad

    @property
    def cantidad_real(self):
        """Retorna el valor actual o actualiza el valor del atributo cantidad_real"""
        return self._cantidad_real

    @property
    def fecha(self):
        """Retorna el valor actual del atributo fecha"""
        return self._fecha

    @property
    def hora_inicio(self):
        """Retorna el valor actual del atributo hora_inicio"""
        return self._hora_inicio

    @property
    def num_error(self):
        """Retorna el valor actual o actualiza el valor del atributo num_error"""
        return self._num_error

    @cantidad.setter
    def cantidad(self, cantidad):
        self._cantidad = cantidad

    @cantidad_real.setter
    def cantidad_real(self, cantidad_real):
        self._cantidad_real = cantidad_real

    @num_error.setter
    def num_error(self, num_error):
        self._num_error = num_error

    def set_param_final(self):
        """Establece parametros finales para medir el tiempo de ejecución del scraper"""
        end = time()
        self._hora_fin = strftime("%H:%M:%S", localtime(end))
        total = end - self._start
        self._tiempo = str(timedelta(seconds=total)).split(".")[0]
        self._productos_por_min = round(self._cantidad / (total / 60), 2)
        self._productos_por_min_real = round(self._cantidad_real / (total / 60), 2)
        log(INFO, f"Errores encontrados: {self._num_error}")
        log(INFO, f"Productos Extraídos: {self._cantidad}")
        log(INFO, f"Hora Fin: {self._hora_fin}")


class ScraperFb:
    """Representa a un bot para hacer web scraping en fb marketplace

    Attributes:
        tiempo (Tiempo): Objeto de la clase Tiempo que maneja información del tiempo de ejecución del scraper
        driver (webdriver.Chrome): Objeto de la clase webdriver que maneja un navegador para hacer web scraping
        wait (WebDriverWait): Objeto de la clase WebDriverWait que maneja el Tiempo de espera durante la ejecución del scraper
        errores (Errores): Objeto de la clase Errores que maneja información de los errores ocurridos durante la ejecución del scraper
        data (Dataset): Objeto de la clase Dataset que maneja información de las publicaciones extraídas por el scraper
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase ScraperFb"""
        log(INFO, "Inicializando scraper")
        self._tiempo = Tiempo()

        # Variable que maneja las opciones de chrome
        chrome_options = ChromeOptions()

        # Configurar nivel de notificacones de chrome
        prefs = {"profile.default_content_setting_values.notifications": 2}
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
        chrome_options.add_argument("--disable-gpu")

        self._driver = Chrome(
            chrome_options=chrome_options,
            service=Service(ChromeDriverManager().install()),
        )
        self._driver.maximize_window()
        self._wait = WebDriverWait(self._driver, 10)
        self._errores = Errores()
        self._data = Dataset()
        log(INFO, f"Hora de inicio: {self._tiempo.hora_inicio}")

    @property
    def data(self):
        """Retorna el valor actual del atributo data"""
        return self._data

    @property
    def errores(self):
        """Retorna el valor actual del atributo errores"""
        return self._errores

    def iniciar_sesion(self, user_name, user_password):
        """Inicia sesión en la página web de facebook usando un usuario y contraseña

        Args:
            user_name (str): Usuario activo de facebook
            user_password (str): Contraseña del usuario activo de facebook
        """
        log(INFO, "Iniciando sesión")
        self._driver.get("https://www.facebook.com/")

        # Localizando los campos de usuario y contraseña
        username = self._wait.until(EC.presence_of_element_located((By.ID, "email")))
        password = self._wait.until(EC.presence_of_element_located((By.ID, "pass")))
        # Limpiando el contenido que existe en los campos de usuario y contraseña
        username.clear()
        password.clear()
        # Mandando valores a los campos de usuario y contraseña
        username.send_keys(user_name)
        password.send_keys(user_password)
        # Dar click en el botón de iniciar sesión
        self._wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[name='login']"))
        ).click()
        # Esperando a que se inicie sesión correctamente
        self._wait.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//a[@class='x1i10hfl x1qjc9v5 xjbqb8w xjqpnuy xa49m3k xqeqjp1 x2hbi6w x13fuv20 xu3j5b3 x1q0q8m5 x26u7qi x972fbf xcfux6l x1qhh985 xm0m39n x9f619 x1ypdohk xdl72j9 x2lah0s xe8uvvx xdj266r x11i5rnm xat24cr x1mh8g0r x2lwn1j xeuugli xexx8yu x4uap5 x18d9i69 xkhd6sd x1n2onr6 x16tdsg8 x1hl2dhg xggy1nq x1ja2u2z x1t137rt x1o1ewxj x3x9cwd x1e5q0jg x13rtm0m x1q0g3np x87ps6o x1lku1pv x1rg5ohu x1a2a7pz x1hc1fzr x1k90msu x6o7n8i xbxq160']",
                )
            )
        )
        log(INFO, "Inicio de sesión con éxito")

    def obtener_publicaciones(self, selector, xpath):
        """Retornar una lista de publicaciones visibles con respecto a una categoría en facebook marketplace

        Args:
            selector (str): Selector a ser usado para localizar las publicaciones
            xpath (str): Ruta de las publicaciones a ser usado por el selector

        Returns:
            list: Lista de publicaciones de Facebook Marketplace
        """
        return self._driver.find_elements(selector, xpath)

    def mapear_datos(self, url):
        """Mapea y extrae los datos de las publicaciones de una categoría

        Args:
            url (str): Link de la página de una categoría en facebook marketplace
        """
        log(INFO, "Accediendo a la URL")
        self._driver.execute_script("window.open('about:blank', 'newtab');")
        self._driver.switch_to.window("newtab")
        self._driver.get(url)

        log(INFO, "Mapeando Publicaciones")
        ropa = self.obtener_publicaciones(
            By.XPATH, '//img[@class="xt7dq6l xl1xv1r x6ikm8r x10wlt62 xh8yej3"]'
        )

        log(INFO, "Creando variables")
        # Enteros que hacen referencia a la fecha en que se postea una publicación y en la que se extrae la información
        fecha_publicacion = fecha_extraccion = int(
            datetime.strptime(self._tiempo.fecha, "%d/%m/%Y").timestamp()
        )
        # Cuenta la cantidad de publicaciones que mapea el scraper
        i = 0
        # Cuenta la cantidad de errores ocurridos durante la ejecución del mapeo del scraper
        e = 0
        enlace = None
        while fecha_publicacion >= fecha_extraccion:
            try:
                log(INFO, f"Scrapeando item {i + 1}")
                # Eliminar de la memoria requests innecesarios
                del self._driver.requests
                # Dar click a la publicación de facebook
                ropa[i].click()
                self._wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//img[@class='x5yr21d xl1xv1r xh8yej3']")
                    )
                )
                # Link de la publicación de facebook
                enlace = sub(
                    r"\?.+", "", self._driver.execute_script("return document.URL")
                )
                for request in self._driver.requests:
                    # Validar si la api es de graphql
                    if not request.response or "graphql" not in request.url:
                        continue

                    # Obtener la respuesta de la api en bytes
                    body = decode(
                        request.response.body,
                        request.response.headers.get("Content-Encoding", "identity"),
                    )
                    # Decodificar la respuesta a utf-8
                    decoded_body = body.decode("utf-8")
                    # Validar si la respuesta decodificada es la deseada
                    if (
                        decoded_body.find(
                            '{"viewer":{"marketplace_product_details_page"'
                        )
                        == -1
                    ):
                        continue

                    # Convertir al formato json la respuesta decodificada anteriormente
                    json_data = loads(decoded_body)

                    # Diccionario que contiene toda la información de la publicación
                    dato = json_data["data"]["viewer"][
                        "marketplace_product_details_page"
                    ]["target"]

                    # Extraer la fecha de publicación
                    fecha_publicacion = dato["creation_time"]

                    log(INFO, f"{dato['marketplace_listing_title']}")
                    self._data.agregar_data(dato, self._tiempo.fecha, enlace)
                    log(INFO, f"Item {i + 1} scrapeado con éxito")
                    break

                # Regresar al inicio donde se encuentran todas las publicaciones de facebook
                self._driver.execute_script("window.history.go(-1)")

            except (
                NoSuchElementException,
                ElementNotInteractableException,
                StaleElementReferenceException,
            ) as error:
                self._errores.agregar_error(error)
                e += 1

            except (
                AttributeError,
                KeyError,
                JSONDecodeError,
                TimeoutException,
            ) as error:
                self._errores.agregar_error(error, enlace)
                self._driver.execute_script("window.history.go(-1)")
                e += 1
            except Exception as error:
                self._errores.agregar_error(error, enlace)
                e += 1
                i += 1
                log(CRITICAL, "Se detuvo inesperadamente el programa")
                log(CRITICAL, f"Causa:\n{error}")
                break

            finally:
                i += 1

                # Verificar si se ha mapeado todas las publicaciones visibles
                if i == len(ropa):
                    # Hacer uso del scroll para obtener más publicaciones
                    self._driver.execute_script(
                        "window.scrollTo(0, document.body.scrollHeight)"
                    )
                    sleep(6)
                    # Mapear las nuevas publicaciones
                    ropa = self.obtener_publicaciones(
                        By.XPATH,
                        '//img[@class="xt7dq6l xl1xv1r x6ikm8r x10wlt62 xh8yej3"]',
                    )
                sleep(2)
                log(
                    INFO,
                    "-------------------------------------------------------------------",
                )

        del self._driver.requests
        # Guardar algunos datos del tiempo de ejecución del scraper
        self._tiempo.cantidad_real = i - e
        self._tiempo.num_error = e
        log(INFO, "Fin de la extraccion")

    def guardar_datos(
        self,
        filetype="Data",
        folder="Data//datos_obtenidos",
        filename="fb_data",
    ):
        """Guarda los datos o errores obtenidos durante la ejecución del scraper

        Args:
            filetype (str, optional): Indica si la información son datos de las publicaciones o errores. Se acepta Data y Error. Defaults to "Data".
            folder (str, optional): Ruta del archivo. Defaults to "Data//datos_obtenidos".
            filename (str, optional): Nombre del archivo. Defaults to "fb_data".
        """
        log(INFO, f"Guardando {filetype}")
        # Comprobando si el valor ingresado para la variable filetype es correcto
        if filetype == "Data":
            # Registrando toda la información de las publicaciones extraídas por el scraper
            dataset = self._data.dataset
        elif filetype == "Error":
            # Registrando toda la información de los errores ocurridos durante la ejecución del scraper
            dataset = self._errores.errores
        else:
            log(
                INFO,
                f"El archivo de tipo {filetype} no está admitido. Solo se aceptan los valores Data y Error",
            )
            log(
                ERROR,
                f"El archivo de tipo {filetype} no se va a guardar por no ser de tipo Data o Error",
            )
            return
        # Crear un dataframe
        df_fb_mkp_ropa = DataFrame(dataset)

        # Comprobando que el dataset contenga información
        if len(df_fb_mkp_ropa) == 0:
            log(
                INFO,
                f"El archivo de tipo {filetype} no se va a guardar por no tener información",
            )
            return

        # Ejecutando diferentes acciones de acuerdo al tipo de información que se va a guardar
        if filetype == "Data":
            # Eliminando la última publicación, porque su fecha de creación es de otro día
            df_fb_mkp_ropa.drop(len(df_fb_mkp_ropa) - 1, axis=0, inplace=True)
            # Registrando la cantidad de información que contiene el dataset
            cantidad = len(df_fb_mkp_ropa)
            self._tiempo.cantidad = cantidad
        else:
            # Registrando la cantidad de errores ocurridos durante la ejecución del scraper
            cantidad = self._tiempo.num_error

        datetime_obj = datetime.strptime(self._tiempo.fecha, "%d/%m/%Y")
        # Generando la ruta donde se va a guardar la información
        filepath = path.join(folder, datetime_obj.strftime("%d-%m-%Y"))
        # Generando el nombre del archivo que va a contener la información
        filename = (
            filename
            + "_"
            + datetime_obj.strftime("%d%m%Y")
            + "_"
            + str(cantidad)
            + ".xlsx"
        )
        # Verificando si la ruta donde se va a guardar la información existe
        if not path.exists(filepath):
            # Creando la ruta donde se va a guardar la información
            makedirs(filepath)
        # Guardando la información en un archivo de tipo excel
        df_fb_mkp_ropa.to_excel(path.join(filepath, filename), index=False)
        log(INFO, f"{filetype} Guardados Correctamente")

    def guardar_tiempos(self, filename, sheet_name):
        """Guarda la información del tiempo de ejecución del scraper

        Args:
            filename (str): Nombre del archivo
            sheet_name (str): Nombre de la hoja de cálculo
        """
        log(INFO, "Guardando tiempos")
        # Guardando los parametros finales del tiempo de ejecución del scraper
        self._tiempo.set_param_final()
        # Variable que indica si el encabezados existe o no en el archivo de excel
        header_exist = True
        # Verificando si el archivo existe o no
        if path.isfile(filename):
            # Leendo el archivo
            tiempos = load_workbook(filename)
            # Comprobando si ya existe un sheet con el nombre indicado en la variable sheet_name
            if sheet_name not in [ws.title for ws in tiempos.worksheets]:
                # Creando un nuevo sheet
                tiempos.create_sheet(sheet_name)
                # Especificar que no existen encabezados en el nuevo sheet
                header_exist = False
        else:
            # Creando un archivo de tipo workbook
            tiempos = Workbook()
            tiempos.worksheets[0].title = sheet_name
            header_exist = False

        # Seleccionar el sheet deseado donde se va a guardar la información
        worksheet = tiempos[sheet_name]

        # Comprobando si el encabezados existe o no
        if not header_exist:
            # Lista que contiene los encabezados a ser insertados
            keys = [
                "Fecha",
                "Hora Inicio",
                "Hora Fin",
                "Cantidad",
                "Cantidad Real",
                "Tiempo Ejecucion (min)",
                "Categorias / Minuto",
                "Categorias / Minuto real",
                "Errores",
            ]
            # Otra forma de indicar los encabezados
            # keys = list(self._tiempo.__dict__.keys())[1:]
            # Insertando los encabezados al sheet
            worksheet.append(keys)
        # Lista que contiene los valores a ser insertados
        values = list(self._tiempo.__dict__.values())[1:]
        # Insertando la información del tiempo al sheet
        worksheet.append(values)
        # Guardar la información en un archivo excel
        tiempos.save(filename)
        # Cerrar el archivo excel
        tiempos.close()
        log(INFO, "Tiempos Guardados Correctamente")


def config_log(log_folder, log_filename, log_file_mode, log_file_encoding):
    """Función que configura los logs para rastrear al programa

    Args:
        log_folder (str): Carpeta donde se va a generar el archivo log
        log_filename (str): Nombre del archivo log a ser generado
        log_file_mode (str): Modo de guardado del archivo
        log_file_encoding (str): Codificación usada para el archivo
    """
    # Mostrar solo los errores de los registros que maneja selenium
    seleniumLogger.setLevel(ERROR)
    # Mostrar solo los errores de los registros que maneja urllib
    urllibLogger.setLevel(ERROR)
    # Mostrar solo los errores de los registros que maneja seleniumwire
    logger = getLogger("seleniumwire")
    logger.setLevel(ERROR)
    environ["WDM_LOG"] = "0"
    # Generando la ruta donde se va a guardar los registros de ejecución
    log_path = path.join(log_folder, CURRENT_DATE.strftime("%d-%m-%Y"))
    # Generando el nombre del archivo que va a contener los registros de ejecución
    log_filename = log_filename + "_" + CURRENT_DATE.strftime("%d%m%Y") + ".log"
    # Verificando si la ruta donde se va a guardar los registros de ejecución existe
    if not path.exists(log_path):
        # Creando la ruta donde se va a guardar los registros de ejecución
        makedirs(log_path)
    # Configuración básica de los logs que maneja este programa
    basicConfig(
        format="%(asctime)s %(message)s",
        level=INFO,
        handlers=[
            StreamHandler(),
            FileHandler(
                path.join(log_path, log_filename), log_file_mode, log_file_encoding
            ),
        ],
    )


def validar_parametros(parametros):
    """Función que valida si los parámetros a usar están definidos

    Args:
        parametros (list): Lista de parámetros

    Returns:
        bool: Indica si los parámetros son validos
    """
    for parametro in parametros:
        # Verifica que el parámetro haya sido definido
        if not parametro:
            log(ERROR, "Parámetros incorrectos")
            return False

    log(INFO, "Parámetros válidos")
    return True


def main():
    try:
        # Formato para el debugger
        config_log("Log", "fb_ropa_log", "w", "utf-8")
        log(INFO, "Configurando Formato Básico del Debugger")

        # Cargar variables de entorno
        log(INFO, "Cargando Variables de entorno")
        load_dotenv()

        # Url de la categoría a scrapear
        url_ropa = getenv("URL_CATEGORY")

        # Parámetros para guardar la data extraída por el scraper
        data_filename = getenv("DATA_FILENAME")
        data_folder = getenv("DATA_FOLDER")

        # Parámetros para guardar la medición de la ejecución del scraper
        filename_tiempos = getenv("FILENAME_TIEMPOS")
        sheet_tiempos = getenv("SHEET_TIEMPOS")

        # Parámetros para guardar los errores durante la ejecución por el scraper
        error_filename = getenv("ERROR_FILENAME")
        error_folder = getenv("ERROR_FOLDER")

        # Parámetros de inicio de sesión
        user = getenv("FB_USERNAME")
        password = getenv("FB_PASSWORD")

        # Validar parámetros
        if not validar_parametros(
            [
                url_ropa,
                data_filename,
                data_folder,
                filename_tiempos,
                sheet_tiempos,
                error_filename,
                error_folder,
                user,
                password,
            ]
        ):
            return

        # Inicializar scrapper
        scraper = ScraperFb()

        # Iniciar sesión
        scraper.iniciar_sesion(user, password)

        # Extracción de datos
        scraper.mapear_datos(url_ropa)

        # Guardando la data extraída por el scraper
        scraper.guardar_datos("Data", data_folder, data_filename)

        # Guardando los errores extraídos por el scraper
        scraper.guardar_datos("Error", error_folder, error_filename)

        # Guardando los tiempos durante la ejecución del scraper
        scraper.guardar_tiempos(filename_tiempos, sheet_tiempos)
        log(INFO, "Programa finalizado")

    except Exception as error:
        log(ERROR, f"Error: {error}")
        log(INFO, "Programa ejecutado con fallos")

    finally:
        try:
            del scraper
        except:
            pass
        # Liberar el archivo log
        shutdown()


if __name__ == "__main__":
    main()
